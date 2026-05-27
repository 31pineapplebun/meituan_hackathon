"""
基于 standard_review_v4_REVIEWED.xlsx 应用复核 → 生成 human_v5_reviewed.csv

输入: 
- human_v4_audited.csv (B.5 自动修订后的版本)
- standard_review_v4_REVIEWED.xlsx (你的人工复核)

输出:
- human_v5_reviewed.csv: 最终标注 v5
- v5_modifications.json: 改动追溯
"""
import csv
import json
from pathlib import Path
from openpyxl import load_workbook
from collections import Counter


def apply_review():
    base = Path("/home/claude/project_v1")
    excel_path = Path("/home/claude/standard_review_v4_REVIEWED.xlsx")
    v4_csv = base / "06_gold_annotation" / "gold_set" / "human_v4_audited.csv"
    v5_csv = base / "06_gold_annotation" / "gold_set" / "human_v5_reviewed.csv"
    log_path = base / "09_pipeline" / "batch_results" / "v5_modifications.json"
    
    # 加载 v4 (作为 base)
    v4_rows = []
    with open(v4_csv, encoding="utf-8-sig") as f:
        v4_rows = list(csv.DictReader(f))
    v4_by_key = {(r["dialogue_id"], r["constraint_id"]): r for r in v4_rows}
    print(f"v4 base 行数: {len(v4_rows)}")
    
    # 加载 Excel 的复核结果
    wb = load_workbook(excel_path)
    ws = wb["待复核标注"]
    
    review_modifications = []
    for r in range(2, ws.max_row + 1):
        dlg_id = ws.cell(row=r, column=2).value
        cid = ws.cell(row=r, column=3).value
        old_verdict = (ws.cell(row=r, column=5).value or "").strip().lower()
        new_verdict = (ws.cell(row=r, column=8).value or "").strip().lower()
        new_evidence = ws.cell(row=r, column=9).value or ""
        
        if not dlg_id or not cid or not new_verdict:
            continue
        
        # 跟v4比较, 看是否真的改动了
        key = (dlg_id, cid)
        v4_row = v4_by_key.get(key)
        if not v4_row:
            continue
        
        v4_verdict = v4_row["verdict"].strip().lower()
        
        # 应用复核: 用 Excel 里的 new_verdict 覆盖 v4
        if new_verdict != v4_verdict or new_evidence:
            # 真的有改动
            review_modifications.append({
                "dialogue_id": dlg_id,
                "constraint_id": cid,
                "v3_verdict": old_verdict,        # Excel "原verdict" 列 = v3
                "v4_verdict": v4_verdict,         # B.5 自动修订后
                "v5_verdict": new_verdict,        # 你复核后
                "v5_evidence": new_evidence[:200],
            })
            
            v4_row["verdict"] = new_verdict
            if new_evidence:
                v4_row["evidence_text"] = new_evidence
            v4_row["notes"] = (v4_row.get("notes") or "") + " [Day9复核]"
    
    # 写 v5
    with open(v5_csv, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=[
            "dialogue_id", "constraint_id", "verdict",
            "evidence_turn", "evidence_text", "confidence", "notes", "rater"
        ])
        writer.writeheader()
        for r in v4_rows:
            r["rater"] = "human_v5_reviewed"
            writer.writerow(r)
    
    # 写日志
    with open(log_path, "w", encoding="utf-8") as f:
        json.dump({
            "total_modifications": len(review_modifications),
            "modifications": review_modifications,
        }, f, ensure_ascii=False, indent=2)
    
    # 统计
    print(f"\n✓ v5 标注完成")
    print(f"  输出: {v5_csv}")
    print(f"  改动数: {len(review_modifications)}")
    
    # 改动方向
    direction = Counter()
    for m in review_modifications:
        direction[(m["v4_verdict"], m["v5_verdict"])] += 1
    print(f"\n  改动方向(v4 → v5):")
    for (o, n), c in direction.most_common():
        print(f"    {o} → {n}: {c}")
    
    return len(review_modifications)


if __name__ == "__main__":
    apply_review()
