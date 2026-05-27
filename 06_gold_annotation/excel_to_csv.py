"""把按指令分sheet的Excel标注转成统一CSV
用于kappa_calc.py的输入

输出格式: dialogue_id, constraint_id, verdict (+其他元数据列)
"""
import argparse
import csv
from openpyxl import load_workbook


SHEETS = ["V1_标注", "V2_标注", "V4_标注", "V5_标注"]


def excel_to_csv(excel_path: str, csv_path: str, rater_name: str = "rater"):
    """转换"""
    wb = load_workbook(excel_path, read_only=True)
    
    rows = []
    for sn in SHEETS:
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        for r in range(2, ws.max_row + 1):
            dialogue_id = ws.cell(row=r, column=1).value
            constraint_id = ws.cell(row=r, column=2).value
            verdict = ws.cell(row=r, column=7).value
            evidence_turn = ws.cell(row=r, column=8).value
            evidence_text = ws.cell(row=r, column=9).value
            confidence = ws.cell(row=r, column=10).value
            notes = ws.cell(row=r, column=11).value
            
            if not dialogue_id or not constraint_id:
                continue
            
            verdict_str = str(verdict).strip().lower() if verdict else ""
            
            rows.append({
                "dialogue_id": dialogue_id,
                "constraint_id": constraint_id,
                "verdict": verdict_str,
                "evidence_turn": str(evidence_turn or ""),
                "evidence_text": str(evidence_text or "")[:500],
                "confidence": str(confidence or "").lower(),
                "notes": str(notes or "")[:200],
                "rater": rater_name,
            })
    
    # 写CSV
    with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
        fieldnames = ["dialogue_id", "constraint_id", "verdict",
                       "evidence_turn", "evidence_text", "confidence", "notes", "rater"]
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    
    # 统计
    from collections import Counter
    verdict_dist = Counter(r["verdict"] for r in rows if r["verdict"])
    print(f"✓ 已写入 {csv_path}")
    print(f"  总行数: {len(rows)}")
    print(f"  非空verdict: {sum(1 for r in rows if r['verdict'])}")
    print(f"  verdict分布: {dict(verdict_dist)}")
    return rows


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", required=True, help="输入Excel(分指令多sheet)")
    parser.add_argument("--csv", required=True, help="输出CSV")
    parser.add_argument("--rater", default="rater", help="标注员名(human/claude/etc)")
    args = parser.parse_args()
    
    excel_to_csv(args.excel, args.csv, args.rater)


if __name__ == "__main__":
    main()
