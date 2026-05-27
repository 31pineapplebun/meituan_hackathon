"""
把生成的对话 JSONL 灌入 Excel 标注模板

用途:
- 输入: dialogues.jsonl (一行一通对话)
- 输入: annotation_template.xlsx (空模板)
- 输出: annotation_filled.xlsx (含对话内容 + 待标注的约束行)

用法:
    python prepare_annotation.py \\
        --dialogues dialogues_output/all_dialogues.jsonl \\
        --template ../06_gold_annotation/annotation_template.xlsx \\
        --output annotation_to_label.xlsx \\
        --constraints ../03_examples/example_2/example_2_atomic.json
"""
import argparse
import json
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


def load_dialogues(path: str) -> list:
    """加载 JSONL"""
    dialogues = []
    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                dialogues.append(json.loads(line))
    return dialogues


def load_constraints(path: str) -> list:
    """加载约束清单"""
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    return data["atomic_constraints"]


def build_filled_workbook(dialogues: list, constraints: list) -> Workbook:
    """创建灌入对话的工作簿"""
    wb = Workbook()
    wb.remove(wb.active)  # 删默认sheet
    
    # === Sheet 1: 对话原文 ===
    ws_dlg = wb.create_sheet("对话原文")
    headers = ["对话ID", "指令", "Persona", "数据源", "Turn", "Role", "内容", "字数"]
    ws_dlg.append(headers)
    
    header_fill = PatternFill("solid", start_color="2E86AB")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col_idx in range(1, len(headers) + 1):
        c = ws_dlg.cell(row=1, column=col_idx)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")
    
    for col_idx, w in {1: 35, 2: 8, 3: 18, 4: 10, 5: 6, 6: 10, 7: 60, 8: 6}.items():
        ws_dlg.column_dimensions[get_column_letter(col_idx)].width = w
    
    asst_fill = PatternFill("solid", start_color="E8F4F8")
    user_fill = PatternFill("solid", start_color="FFF4E6")
    
    for dlg in dialogues:
        for t in dlg["turns"]:
            row = [
                dlg["dialogue_id"],
                dlg["instruction_name"],
                dlg["persona_id"],
                dlg.get("sample_source", ""),  # 新增数据源列
                t["turn"],
                t["role"],
                t["content"],
                len(t["content"]),
            ]
            ws_dlg.append(row)
            r = ws_dlg.max_row
            fill = asst_fill if t["role"] == "assistant" else user_fill
            for c in range(1, len(headers) + 1):
                ws_dlg.cell(row=r, column=c).fill = fill
                ws_dlg.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical="top")
    
    ws_dlg.freeze_panes = "A2"
    
    # === Sheet 2: 标注 ===
    ws_ann = wb.create_sheet("标注")
    ann_headers = ["对话ID", "约束ID", "约束名称", "维度", "is_critical", 
                    "verdict", "evidence_turn", "evidence_text", "confidence", "notes"]
    ws_ann.append(ann_headers)
    
    for col_idx in range(1, len(ann_headers) + 1):
        c = ws_ann.cell(row=1, column=col_idx)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    widths = {1: 35, 2: 12, 3: 32, 4: 22, 5: 12, 6: 12, 7: 14, 8: 40, 9: 12, 10: 30}
    for col_idx, w in widths.items():
        ws_ann.column_dimensions[get_column_letter(col_idx)].width = w
    
    dim_name_map = {
        "D1_flow_compliance": "D1 流程遵循度",
        "D2_task_completion": "D2 任务完成度",
        "D3_constraint_compliance": "D3 约束遵循度",
        "D4_knowledge_accuracy": "D4 知识准确性",
        "D5_dialogue_quality": "D5 对话质量"
    }
    
    crit_fill = PatternFill("solid", start_color="FFF2CC")
    row = 2
    for dlg in dialogues:
        for c in constraints:
            ws_ann.cell(row=row, column=1, value=dlg["dialogue_id"])
            ws_ann.cell(row=row, column=2, value=c["id"])
            ws_ann.cell(row=row, column=3, value=c["name"])
            ws_ann.cell(row=row, column=4, value=dim_name_map.get(c["scoring_dimension"], c["scoring_dimension"]))
            is_crit = c.get("is_critical", False)
            ws_ann.cell(row=row, column=5, value="是" if is_crit else "")
            if is_crit:
                for col in range(1, 11):
                    ws_ann.cell(row=row, column=col).fill = crit_fill
            row += 1
    
    # 下拉数据验证
    dv_verdict = DataValidation(type="list", formula1='"pass,fail,na,review"', allow_blank=True)
    dv_verdict.add(f"F2:F{row-1}")
    ws_ann.add_data_validation(dv_verdict)
    
    dv_conf = DataValidation(type="list", formula1='"high,medium,low"', allow_blank=True)
    dv_conf.add(f"I2:I{row-1}")
    ws_ann.add_data_validation(dv_conf)
    
    ws_ann.freeze_panes = "A2"
    
    # === Sheet 3: 进度统计 ===
    ws_stat = wb.create_sheet("进度统计")
    ws_stat.append(["对话数", "约束数", "标注总单元", "已完成", "完成率"])
    for col_idx in range(1, 6):
        c = ws_stat.cell(row=1, column=col_idx)
        c.fill = header_fill
        c.font = header_font
    
    n_dlg = len(dialogues)
    n_cs = len(constraints)
    total_cells = n_dlg * n_cs
    ws_stat.append([n_dlg, n_cs, total_cells, "=COUNTIF(标注!F:F,\"<>\")-1", 
                     f"=ROUND(D2/{total_cells}*100,1)&\"%\""])
    
    for col_idx, w in {1: 12, 2: 12, 3: 14, 4: 12, 5: 12}.items():
        ws_stat.column_dimensions[get_column_letter(col_idx)].width = w
    
    return wb


def main():
    parser = argparse.ArgumentParser(description="灌入对话到标注模板")
    parser.add_argument("--dialogues", required=True, help="JSONL 对话文件")
    parser.add_argument("--constraints", required=True, help="约束清单 JSON")
    parser.add_argument("--output", default="annotation_to_label.xlsx", help="输出 Excel")
    args = parser.parse_args()
    
    print(f"加载对话: {args.dialogues}")
    dialogues = load_dialogues(args.dialogues)
    print(f"  共 {len(dialogues)} 通对话")
    
    print(f"加载约束: {args.constraints}")
    constraints = load_constraints(args.constraints)
    print(f"  共 {len(constraints)} 条约束")
    
    print(f"\n构建标注 Excel...")
    wb = build_filled_workbook(dialogues, constraints)
    wb.save(args.output)
    
    print(f"\n✓ 已生成: {args.output}")
    print(f"  对话原文 Sheet: {sum(len(d['turns']) for d in dialogues)} 行")
    print(f"  标注 Sheet: {len(dialogues) * len(constraints)} 行待标")
    print(f"  进度统计 Sheet: 自动计算完成率")


if __name__ == "__main__":
    main()
